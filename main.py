import requests
import json
from datetime import datetime, timedelta
import ephem
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from os import system, name

#=============================================================================================#

def NOAA_API_call(station_id, date): # modified to take the day after as well
    NOAA_url = "https://api.tidesandcurrents.noaa.gov/api/prod/datagetter" # will always be this URL

    next_date = getNextDay(date) # calling api for 2 days

    params = { 
        "product": "high_low",  # Changed to get extremes
        "station": station_id,
        "begin_date": date,  # Fetch today's data
        "end_date": next_date, # (required for high_low product)
        "datum": "MLLW",    # grab data in MLLW in case tide station doesn't have available data for NAVD
        "units": "english",
        "time_zone": "GMT",     # Must be uppercase "GMT"
        "format": "json" } 

    response = requests.get(NOAA_url, params) # using requests library to perform a GET request to API with specified above params

    return response

#=============================================================================================#

def recordJSONResponse(tide_values, response): # takes station_id, date, and a reference to tide_values
# function will modify tide_values with HH value for given date and return nothing
    
    if response.status_code == 200: # Check if request was successful
        try:
            WLdata = response.json() # convert response into variable WLdata

            with open("data.txt", "a") as file: # appending json response into text file
                json.dump(WLdata, file, indent=4)

            if "data" in WLdata: # appends Higher high tide data to tide_values
                highestEntry = 0
                for entry in WLdata["data"]: # for each entry in "data" response
                    currentEntry = float(entry["v"])
                    if currentEntry > highestEntry:
                        highestEntry = currentEntry
                    
                tide_values.append(highestEntry)
                print("Appended:", highestEntry) # printing appended value

        except json.JSONDecodeError:
            print("Error: Failed to decode JSON response")
    else:
        print(f"Error: API request failed with status code {response.status_code}")
        print(response.text)

#=============================================================================================#

def calc_median(nums): # simple function to calculate the median of a list
    nums.sort()
    n = len(nums)
    m = n // 2
    if m % 2 == 0:
        return ((nums[m-1] + nums[m]) / 2)
    elif m % 2 == 1:
        return nums[m]

def calc_mean(nums):
    sum = 0
    for i in nums:
        sum += i
    mean = sum/len(nums)

    return mean

#=============================================================================================#
    
def determineSpringMoonDates(start_date_str, end_date_str): # will determine all the days that the spring(full/new) moon are and modify date_list accordingly
    # Convert input strings to datetime objects
    start_date = datetime.strptime(start_date_str, "%Y%m%d")
    end_date = datetime.strptime(end_date_str, "%Y%m%d")
    
    # Initialize list for moon dates
    date_list = []
    
    # Create observer (timezone set to UTC/GMT)
    observer = ephem.Observer()
    observer.elevation = -6  # Makes times more accurate
    
    current_date = start_date
    while current_date <= end_date:
        # Set observer time to noon UTC for current date
        observer.date = current_date.strftime('%Y/%m/%d 12:00:00')
        
        # Calculate moon phase (0 = new moon, 1 = full moon)
        moon = ephem.Moon(observer.date)
        moon_phase = moon.moon_phase
        
        # Check for new moon (0) or full moon (1)
        if moon_phase < 0.01:  # New moon threshold
            date_list.append(current_date.strftime("%Y%m%d"))
            current_date += timedelta(days=5)

        elif moon_phase > 0.99:  # Full moon threshold
            date_list.append(current_date.strftime("%Y%m%d"))
            current_date += timedelta(days=5)
        
        current_date += timedelta(days=1)
    
    return date_list

def getNextDay(current_date): # takes current date as string and gets the next day and returns it as string
    current_date = datetime.strptime(current_date, "%Y%m%d")
    next_day = current_date + timedelta(days=1)
    return str(next_day.strftime("%Y%m%d"))

#=============================================================================================#
# user input (temporary for date range)

def initializeSet():
    noaaIDSet = set()
    with open("noaa_stations.txt", "r") as file:
        for line in file:
            noaaIDSet.add(line.strip())
    return noaaIDSet

def getStationID(noaaIDSet):
    while (True):
        userStationID = input("Enter NOAA Station ID: ")
        if userStationID in noaaIDSet:
            break
        else:
            print("Invalid Station ID. Using Station ID for The Battery NY.")
            return "8518750"
    return userStationID

def grabStartDate():
    while (True):
        start_date = input("\nEnter start date in YYYYMMDD format:")
        if len(start_date) != 8:
            print("Invalid input.")
        else:
            break
    return start_date

def grabEndDate():
    while (True):
        end_date = input("\nEnter end date in YYYYMMDD format:")
        if len(end_date) != 8:
            print("Invalid input.")
        else:
            break
    return end_date

#=============================================================================================#

def exportToExcel(date_list, tide_values, median, mean):
    
    # Create DataFrame with main data
    tidevaluedf = pd.DataFrame(list(zip(date_list, tide_values)),
                        columns=['Date', 'HH Tide Values (MLLW)'])
    
    # Write to Excel
    tidevaluedf.to_excel('Spring_High_Data.xlsx', index=False)
    
    # Load workbook to add statistics
    wb = load_workbook('Spring_High_Data.xlsx')
    ws = wb.active
    
    # Calculate position for statistics
    stats_row = len(date_list) + 3  # Leave 2 blank rows after data
    
    # Add median
    ws.cell(row=stats_row, column=1, value="Median:")
    ws.cell(row=stats_row, column=2, value=median)
    
    # Add mean
    ws.cell(row=stats_row+1, column=1, value="mean:")
    ws.cell(row=stats_row+1, column=2, value=mean)
    
    # Format columns
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = 25
        
        # Apply number formatting to values column
        if column_letter == 'B':
            for cell in column:
                if isinstance(cell.value, float):
                    cell.number_format = '0.000'
    
    # Format statistics as bold
    for row in [stats_row, stats_row+1]:
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = '0.000'
    
    wb.save('Spring_High_Data.xlsx')
    print("Spring_High_Data.xlsx updated.")

#=============================================================================================#

def printTideValues(date_list, tide_values, median, mean): # function to print tide values pretty in terminal

    # splitting up first and last value in date list and adding slashes
    start_date = f"{date_list[0][:4]}/{date_list[0][4:6]}/{date_list[0][6:8]}"
    end_date = f"{date_list[-1][:4]}/{date_list[-1][4:6]}/{date_list[-1][6:8]}"

    print(f"\nAll Tide values (MLLW) from {start_date} to {end_date} in chronological order:")
    print("Date - - - - - | Tide Value - - - - - |")
    i = 0 # setting counter for tracking tide_values in for loop

    for date in date_list:
        formatted_date = f"{date[:4]}/{date[4:6]}/{date[6:8]}" # formatting date same as above
        print(formatted_date.ljust(16, ' '), end=" ")
        if i < (len(tide_values)):
            print(tide_values[i])
        i += 1

    print("\nMedian : %.2f" % median)
    print("mean : %.2f" % mean)

#=============================================================================================#

def wipeDataTxt():
    with open("data.txt", "w") as file:
        file.write(" ")

#=============================================================================================#

def printMainMenu(station_id, start_date, end_date): # main menu print function

    if name == 'nt': # clears screen for windows or mac/linux terminal
        _ = system('cls')
    else:
        _ = system('clear')

    print("~~ Spring High Water Calculator ~~")
    print("Real-time and historical data obtained using NOAA Co-ops API")
    print(f"Current Station: {station_id}") # print out N/A in case of no station
    print(f"Current Date Range: {start_date} to {end_date} ") # print out N/A in case of no station

    print("\n[0] - Update NOAA tide station ID") 
    print("[1] - Update start and end date")
    print("[2] - Perform NOAA API call")
    print("[3] - Display data for last API call")
    print("[4] - Export data to excel")
    print("[-1] - Exit")

#=============================================================================================#

def main():
    #initializing variables
    tide_values = [] # list to store all higher high tide values
    noaaIDSet = initializeSet()

    # getting values from user
    station_id = str(getStationID(noaaIDSet))
    start_date = str(grabStartDate())
    end_date = str(grabEndDate())

    # determining date list based on user-entered range
    date_list = determineSpringMoonDates(start_date, end_date)


    # wiping old json response data
    wipeDataTxt()

    # appending all HH tide values to list according to all dates in date_list
    for date in date_list:
        response = NOAA_API_call(station_id, date)
        recordJSONResponse(tide_values, response)
    
    #print(tide_values) # checking what tide_values list looks like

    # calculating median and mean
    copyTideList = list(tide_values)
    median = calc_median(copyTideList)
    mean = calc_mean(copyTideList)

    # printing to terminal and updating excel sheet
    printTideValues(date_list, tide_values, median, mean)
    exportToExcel(date_list, tide_values, median, mean)

    # testing
    stopInput = input("\n[0] - Continue")
    system('cls')
    printMainMenu(station_id, start_date, end_date)
    stopInput = input("")

#=============================================================================================#

if __name__ == "__main__":
    main()